import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class FileManager:

    file_name = "job_data.xlsx"

    @classmethod
    def save_file(cls, job_data):

        df = pd.json_normalize(job_data)

        if os.path.exists(cls.file_name):
            existing_data = pd.read_excel(cls.file_name)
            updated_data = pd.concat([existing_data, df], ignore_index=True)
        else:
            updated_data = df

        updated_data.to_excel(cls.file_name, index=False)

        cls.wrap_cell()

    @classmethod
    def wrap_cell(cls):
        wb = load_workbook(cls.file_name)
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        wb.save(cls.file_name)
